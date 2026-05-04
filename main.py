# -*- coding: utf-8 -*-
"""
随机抽选工具 - App14345
最终版：多次抽取弹窗自动关闭改为独立复选框，默认启用0.5秒。
"""

import sys
import os
import json
import shutil
import secrets
import random
import string
import time
import tempfile
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from functools import partial

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QPushButton, QLabel, QLineEdit, QSpinBox, QMessageBox, QFileDialog,
    QDialog, QDialogButtonBox, QScrollArea, QFrame, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QSystemTrayIcon, QMenu, QStyle, QTextEdit, QSplitter, QGroupBox,
    QAbstractItemView, QRadioButton, QButtonGroup, QDoubleSpinBox
)
from PyQt5.QtCore import (
    Qt, QTimer, QSize, QPoint, QRect, QByteArray, QPropertyAnimation,
    QEasingCurve, QUrl
)
from PyQt5.QtNetwork import QLocalServer, QLocalSocket
from PyQt5.QtGui import (
    QIcon, QPixmap, QColor, QBrush, QPainter, QImage, QFont, QDesktopServices
)

# matplotlib 配置中文
import matplotlib
matplotlib.use('Qt5Agg')
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# openpyxl
try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==================== 全局路径 ====================
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()
CONFIG_DIR = os.path.join(BASE_DIR, "config")
PERSON_DIR = os.path.join(BASE_DIR, "person")
PHOTO_DIR = os.path.join(PERSON_DIR, "photo")
LOG_DIR = os.path.join(BASE_DIR, "日志")
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
ICON_PATH = os.path.join(BASE_DIR, "app.ico")

for d in [CONFIG_DIR, PERSON_DIR, PHOTO_DIR, LOG_DIR, TEMPLATE_DIR]:
    os.makedirs(d, exist_ok=True)

SETTINGS_FILE = os.path.join(CONFIG_DIR, "settings.json")
PERSONS_FILE = os.path.join(PERSON_DIR, "persons.json")
STATS_FILE = os.path.join(CONFIG_DIR, "stats.json")

def ensure_default_icon():
    if not os.path.exists(ICON_PATH):
        img = Image.new('RGB', (64, 64), color='red')
        img.save(ICON_PATH)

def generate_import_template():
    """生成带居中文字和格式的导入模板"""
    if not OPENPYXL_AVAILABLE:
        return None
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "人员导入"
    headers = ["序号", "绰号*", "姓名", "头像"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([1, "张三", "张三山", "a.jpg"])
    ws.append([2, "李四", "李四明", "b.jpg"])
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    template_path = os.path.join(TEMPLATE_DIR, "人员导入模板.xlsx")
    wb.save(template_path)
    return template_path

# ==================== 日志系统 ====================
class Logger:
    def __init__(self):
        now = datetime.now()
        random_suffix = ''.join(random.choices(string.digits, k=4))
        filename = f"{now.strftime('%Y%m%d_%H%M%S')}_{random_suffix}.log"
        self.log_file = os.path.join(LOG_DIR, filename)
        self._write(f"日志初始化 | 时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')}")

    def _write(self, text):
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(text + '\n')

    def info(self, msg):
        self._write(f"[INFO] {msg}")

    def log_draw(self, data):
        self._write(f"[DRAW] {json.dumps(data, ensure_ascii=False)}")

    def log_raw(self, text):
        self._write(text)

# ==================== 人员模型 ====================
class Person:
    def __init__(self, pid, nickname, realname="", photo="", participate=True):
        self.id = pid
        self.nickname = nickname
        self.realname = realname
        self.photo = photo
        self.participate = participate

    def to_dict(self):
        return {
            "id": self.id,
            "nickname": self.nickname,
            "realname": self.realname,
            "photo": self.photo,
            "participate": self.participate
        }

    @staticmethod
    def from_dict(d):
        return Person(d["id"], d["nickname"], d.get("realname", ""), d.get("photo", ""), d.get("participate", True))

class PersonManager:
    def __init__(self):
        self.persons = []
        self.load()

    def load(self):
        if os.path.exists(PERSONS_FILE):
            with open(PERSONS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.persons = [Person.from_dict(item) for item in data]
        else:
            self.generate_sample_persons()
            self.save()

    def save(self):
        with open(PERSONS_FILE, 'w', encoding='utf-8') as f:
            json.dump([p.to_dict() for p in self.persons], f, ensure_ascii=False, indent=2)

    def generate_sample_persons(self):
        import random as rnd
        first_names = ["张", "王", "李", "赵", "陈", "林", "黄", "刘", "周", "吴"]
        last_names = ["伟", "芳", "娜", "敏", "静", "涛", "军", "强", "鹏", "杰"]
        nick_suffix = ["星", "侠", "司机", "新", "霸", "货", "猫", "龙", "客", "女"]
        count = rnd.randint(15, 20)
        for i in range(count):
            nickname = f"{rnd.choice(nick_suffix)}_{i+1}"
            realname = f"{rnd.choice(first_names)}{rnd.choice(last_names)}"
            person = Person(str(i+1), nickname, realname, "", True)
            self.persons.append(person)

    def get_participants(self):
        return [p.id for p in self.persons if p.participate]

    def get_person_by_id(self, pid):
        for p in self.persons:
            if p.id == pid:
                return p
        return None

    def add_person(self, nickname, realname, photo_path=None):
        new_id = str(int(max([int(p.id) for p in self.persons], default=0)) + 1)
        photo_filename = ""
        if photo_path and os.path.exists(photo_path):
            ext = os.path.splitext(photo_path)[1]
            photo_filename = f"{new_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
            dest = os.path.join(PHOTO_DIR, photo_filename)
            shutil.copy2(photo_path, dest)
        person = Person(new_id, nickname, realname, photo_filename, True)
        self.persons.append(person)
        self.save()
        return person

    def update_person(self, pid, nickname, realname, photo_path=None):
        person = self.get_person_by_id(pid)
        if person:
            person.nickname = nickname
            person.realname = realname
            if photo_path and os.path.exists(photo_path):
                if person.photo:
                    old_path = os.path.join(PHOTO_DIR, person.photo)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                ext = os.path.splitext(photo_path)[1]
                new_filename = f"{pid}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                dest = os.path.join(PHOTO_DIR, new_filename)
                shutil.copy2(photo_path, dest)
                person.photo = new_filename
            self.save()
            return True
        return False

    def delete_person(self, pid):
        person = self.get_person_by_id(pid)
        if person:
            if person.photo:
                photo_path = os.path.join(PHOTO_DIR, person.photo)
                if os.path.exists(photo_path):
                    os.remove(photo_path)
            self.persons = [p for p in self.persons if p.id != pid]
            self.save()
            return True
        return False

    def set_participate(self, pid, participate):
        person = self.get_person_by_id(pid)
        if person:
            person.participate = participate
            self.save()
            return True
        return False

    def import_from_excel(self, file_path, photo_mode='none', temp_photo_dir=None):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(None, "错误", "未安装 openpyxl 模块")
            return 0
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"读取Excel失败：{e}")
            return 0
        if "人员导入" not in wb.sheetnames:
            QMessageBox.critical(None, "错误", "Excel文件中缺少「人员导入」工作表")
            return 0
        ws = wb["人员导入"]
        rows = list(ws.iter_rows(min_row=2, values_only=False))
        imported_cnt = 0
        embedded_images = {}
        if hasattr(ws, '_images'):
            for img in ws._images:
                try:
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        col = img.anchor._from.col + 1
                        row = img.anchor._from.row + 1
                        img_data = img._data()
                        fd, temp_path = tempfile.mkstemp(suffix='.png', prefix='excel_img_')
                        os.close(fd)
                        with open(temp_path, 'wb') as f:
                            f.write(img_data)
                        embedded_images[(row, col)] = temp_path
                except Exception as e:
                    print(f"提取内嵌图片失败: {e}")

        for row in rows:
            if not row or len(row) < 4:
                continue
            seq_cell, nickname_cell, realname_cell, avatar_cell = row[0], row[1], row[2], row[3]
            seq = seq_cell.value
            nickname = nickname_cell.value
            realname = realname_cell.value
            avatar = avatar_cell.value if avatar_cell.value else ""
            if not nickname:
                continue
            nickname = str(nickname).strip()
            realname = str(realname).strip() if realname else ""
            avatar_file = str(avatar).strip() if avatar else ""
            photo_path = None
            for (r, c), path in embedded_images.items():
                if r == seq_cell.row and c == 4:
                    photo_path = path
                    break
            if photo_path is None and photo_mode == 'direct' and avatar_file:
                if os.path.exists(avatar_file):
                    photo_path = avatar_file
                elif os.path.exists(os.path.join(os.path.dirname(file_path), avatar_file)):
                    photo_path = os.path.join(os.path.dirname(file_path), avatar_file)
            elif photo_mode == 'batch' and temp_photo_dir and seq is not None:
                seq_num = int(seq)
                for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                    for prefix in ['img ', '人员img ']:
                        test_name = f"{prefix}({seq_num}){ext}"
                        test_path = os.path.join(temp_photo_dir, test_name)
                        if os.path.exists(test_path):
                            photo_path = test_path
                            break
                    if photo_path:
                        break
            if photo_path:
                ext = os.path.splitext(photo_path)[1]
                new_id = str(len(self.persons) + 1)
                new_filename = f"{new_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                dest = os.path.join(PHOTO_DIR, new_filename)
                shutil.copy2(photo_path, dest)
                avatar_file = new_filename
                if photo_path.startswith(tempfile.gettempdir()):
                    try:
                        os.remove(photo_path)
                    except:
                        pass
            else:
                avatar_file = ""
            new_id = str(len(self.persons) + 1)
            person = Person(new_id, nickname, realname, avatar_file, True)
            self.persons.append(person)
            imported_cnt += 1
        self.save()
        return imported_cnt

# ==================== 随机抽取核心 ====================
class RandomPicker:
    def __init__(self, logger, ui_callback=None, person_manager=None):
        self.logger = logger
        self.ui_callback = ui_callback
        self.pm = person_manager

    def get_valid_microsecond(self, max_retry=50):
        for _ in range(max_retry):
            ms = datetime.now().microsecond
            if ms != 0:
                return ms
            time.sleep(0.0001)
            QApplication.processEvents()
        return 1

    def pick(self, participant_ids, k=1, with_replacement=False):
        if not participant_ids:
            return []
        if k > len(participant_ids) and not with_replacement:
            k = len(participant_ids)

        N = len(participant_ids)
        shuffled = participant_ids.copy()
        secrets.SystemRandom().shuffle(shuffled)
        seq_to_id = {idx+1: pid for idx, pid in enumerate(shuffled)}
        max_digits = len(str(N))
        mapping_lines = []
        line = []
        for seq in range(1, N+1):
            pid = seq_to_id[seq]
            person = self.pm.get_person_by_id(pid)
            name = person.realname if person.realname else person.nickname
            line.append(f"{seq:>{max_digits}}→{name}")
            if len(line) == 5 or seq == N:
                mapping_lines.append(", ".join(line) + ";")
                line = []
        mapping_str = "\n" + "\n".join(mapping_lines)
        if self.ui_callback:
            self.ui_callback(f"本次映射顺序: {mapping_str}", timestamp=False, deferrable=False)
        self.logger.log_draw({"step": "seq_mapping", "mapping": {pid: i+1 for i, pid in enumerate(shuffled)}})

        total_valid = 999999
        block_size = total_valid // N
        remainder = total_valid % N
        max_valid_ms = total_valid - remainder

        chosen = []
        remaining_ids = participant_ids.copy()
        current_seq_to_id = seq_to_id.copy()
        current_N = N
        current_max_valid = max_valid_ms
        current_block = block_size

        for step in range(k):
            if not remaining_ids:
                break
            if not with_replacement and step > 0:
                remaining_ids = [pid for pid in remaining_ids if pid not in chosen]
                current_N = len(remaining_ids)
                if current_N == 0:
                    break
                shuffled2 = remaining_ids.copy()
                secrets.SystemRandom().shuffle(shuffled2)
                current_seq_to_id = {idx+1: pid for idx, pid in enumerate(shuffled2)}
                current_block = total_valid // current_N
                current_rem = total_valid % current_N
                current_max_valid = total_valid - current_rem
                if self.ui_callback:
                    self.ui_callback(f"🔄 剩余 {current_N} 人，重新映射", deferrable=False)

            while True:
                ms = self.get_valid_microsecond()
                if ms <= current_max_valid:
                    break
                self.logger.log_draw({"step": "discard_ms", "ms": ms})
                if self.ui_callback:
                    self.ui_callback(f"🎲 舍弃无效微秒: {ms} (>{current_max_valid})", timestamp=True)
            divisor = current_N
            remainder_div = ms % divisor
            quotient = ms // divisor
            seq = divisor if remainder_div == 0 else remainder_div
            seq = max(1, min(seq, divisor))
            chosen_id = current_seq_to_id[seq]
            person = self.pm.get_person_by_id(chosen_id)
            name = person.realname if person.realname else person.nickname
            log_text = f"🎲 微秒={ms} → {ms}/{divisor} = \n↓↓↓-------↓-------↓-------↓-------↓↓↓---\n{quotient} ··· {remainder_div} → 本轮{seq}={name} → 抽中: {name}"
            if self.ui_callback:
                self.ui_callback(log_text, timestamp=True)
            self.logger.log_draw({"step": f"draw_{step+1}", "ms": ms, "seq": seq, "chosen_id": chosen_id})
            chosen.append(chosen_id)

        return chosen

# ==================== 裁剪对话框 ====================
class CropDialog(QDialog):
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.image_path = image_path
        self.original_pixmap = QPixmap(image_path)
        if self.original_pixmap.isNull():
            QMessageBox.critical(self, "错误", "无法加载图片")
            self.reject()
            return
        self.init_ui()
        self.setWindowTitle("裁剪头像")
        self.resize(800, 600)
        self.scale_factor = 1.0
        self.rubber_band = None
        self.origin_point = QPoint()
        self.crop_rect = QRect()
        self.update_scaled_pixmap()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.scroll_area.setWidget(self.image_label)
        layout.addWidget(self.scroll_area)
        btn_layout = QHBoxLayout()
        self.btn_cancel = QPushButton("取消")
        self.btn_save = QPushButton("保存")
        btn_layout.addWidget(self.btn_cancel)
        btn_layout.addWidget(self.btn_save)
        layout.addLayout(btn_layout)
        self.btn_save.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)

        self.image_label.setMouseTracking(True)
        self.image_label.mousePressEvent = self.on_mouse_press
        self.image_label.mouseMoveEvent = self.on_mouse_move
        self.image_label.mouseReleaseEvent = self.on_mouse_release

    def update_scaled_pixmap(self):
        if self.original_pixmap.isNull():
            return
        label_size = self.scroll_area.viewport().size()
        if label_size.width() <= 0 or label_size.height() <= 0:
            return
        self.scaled_pixmap = self.original_pixmap.scaled(
            label_size, Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        self.scale_factor = self.scaled_pixmap.width() / self.original_pixmap.width()
        self.image_label.setPixmap(self.scaled_pixmap)
        self.image_label.resize(self.scaled_pixmap.size())

    def resizeEvent(self, event):
        self.update_scaled_pixmap()
        super().resizeEvent(event)

    def on_mouse_press(self, event):
        self.origin_point = event.pos()
        if not self.rubber_band:
            self.rubber_band = QRubberBand(QRubberBand.Rectangle, self.image_label)
        self.rubber_band.setGeometry(QRect(self.origin_point, QSize()))
        self.rubber_band.show()

    def on_mouse_move(self, event):
        if self.rubber_band:
            self.rubber_band.setGeometry(QRect(self.origin_point, event.pos()).normalized())

    def on_mouse_release(self, event):
        if self.rubber_band:
            self.crop_rect = self.rubber_band.geometry()
            self.rubber_band.hide()
            msg = QMessageBox.question(self, "确认裁剪", "是否使用此区域？", QMessageBox.Yes | QMessageBox.No)
            if msg == QMessageBox.Yes:
                self.accept()
            else:
                self.crop_rect = QRect()

    def get_cropped_pixmap(self):
        if not self.crop_rect.isNull() and not self.crop_rect.isEmpty():
            x = int(self.crop_rect.x() / self.scale_factor)
            y = int(self.crop_rect.y() / self.scale_factor)
            w = int(self.crop_rect.width() / self.scale_factor)
            h = int(self.crop_rect.height() / self.scale_factor)
            x = max(0, min(x, self.original_pixmap.width() - 1))
            y = max(0, min(y, self.original_pixmap.height() - 1))
            w = max(1, min(w, self.original_pixmap.width() - x))
            h = max(1, min(h, self.original_pixmap.height() - y))
            return self.original_pixmap.copy(x, y, w, h)
        else:
            size = min(self.original_pixmap.width(), self.original_pixmap.height())
            x = (self.original_pixmap.width() - size) // 2
            y = (self.original_pixmap.height() - size) // 2
            return self.original_pixmap.copy(x, y, size, size)

    def save_cropped(self, save_path):
        cropped = self.get_cropped_pixmap()
        cropped = cropped.scaled(256, 256, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        cropped.save(save_path, "PNG")
        return True

# ==================== 人员管理对话框 ====================
class PersonManageDialog(QDialog):
    def __init__(self, person_manager, parent=None):
        super().__init__(parent)
        self.pm = person_manager
        self.setWindowTitle("人员维护")
        self.setMinimumSize(850, 600)
        self.setWindowIcon(QIcon(ICON_PATH) if os.path.exists(ICON_PATH) else QIcon())
        self.init_ui()
        self.refresh_list()

    def init_ui(self):
        layout = QVBoxLayout(self)
        toolbar = QHBoxLayout()
        self.btn_import = QPushButton("批量导入")
        self.btn_import.clicked.connect(self.batch_import)
        toolbar.addWidget(self.btn_import)
        self.btn_download_template = QPushButton("下载模板")
        self.btn_download_template.clicked.connect(self.download_template)
        toolbar.addWidget(self.btn_download_template)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        search_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索昵称/真实姓名")
        self.search_edit.textChanged.connect(self.refresh_list)
        search_layout.addWidget(self.search_edit)
        layout.addLayout(search_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["昵称", "真实姓名", "参与抽取", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("新增人员")
        self.btn_add.clicked.connect(self.add_person)
        btn_layout.addWidget(self.btn_add)
        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(self.btn_close)
        layout.addLayout(btn_layout)

    def refresh_list(self):
        self.table.setRowCount(0)
        keyword = self.search_edit.text().strip()
        for person in self.pm.persons:
            if keyword and keyword not in person.nickname and keyword not in person.realname:
                continue
            row = self.table.rowCount()
            self.table.insertRow(row)
            nick_item = QTableWidgetItem(person.nickname)
            nick_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, nick_item)
            real_item = QTableWidgetItem(person.realname)
            real_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 1, real_item)
            cb_widget = QWidget()
            cb_layout = QHBoxLayout(cb_widget)
            cb_layout.setAlignment(Qt.AlignCenter)
            cb = QCheckBox()
            cb.setChecked(person.participate)
            cb.stateChanged.connect(partial(self.on_participate_changed, person.id))
            cb_layout.addWidget(cb)
            cb_layout.setContentsMargins(0,0,0,0)
            self.table.setCellWidget(row, 2, cb_widget)
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(0,0,0,0)
            btn_edit = QPushButton("编辑")
            btn_edit.clicked.connect(partial(self.edit_person, person.id))
            btn_del = QPushButton("删除")
            btn_del.clicked.connect(partial(self.delete_person, person.id))
            btn_layout.addWidget(btn_edit)
            btn_layout.addWidget(btn_del)
            self.table.setCellWidget(row, 3, btn_widget)

    def on_participate_changed(self, pid, state):
        self.pm.set_participate(pid, state == Qt.Checked)

    def add_person(self):
        dialog = PersonEditDialog(self.pm, None, self)
        if dialog.exec_():
            self.refresh_list()

    def edit_person(self, pid):
        person = self.pm.get_person_by_id(pid)
        if person:
            dialog = PersonEditDialog(self.pm, person, self)
            if dialog.exec_():
                self.refresh_list()

    def delete_person(self, pid):
        if QMessageBox.question(self, "确认", "删除人员将同时删除头像，是否继续？") == QMessageBox.Yes:
            self.pm.delete_person(pid)
            self.refresh_list()

    def download_template(self):
        template_path = generate_import_template()
        if template_path:
            QDesktopServices.openUrl(QUrl.fromLocalFile(TEMPLATE_DIR))
            QMessageBox.information(self, "提示", f"模板已生成，保存在：{template_path}")
        else:
            QMessageBox.warning(self, "警告", "请先安装 openpyxl 模块以生成模板。")

    def batch_import(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "错误", "请先安装 openpyxl 模块。")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("批量导入人员")
        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("请选择导入模式："))
        btn_direct = QPushButton("直接导入（带照片，支持单元格内嵌图片）")
        btn_no_photo = QPushButton("先导入无照片名单（稍后匹配头像）")
        layout.addWidget(btn_direct)
        layout.addWidget(btn_no_photo)
        btn_cancel = QPushButton("取消")
        layout.addWidget(btn_cancel)
        result = [None]
        def on_direct():
            result[0] = "direct"
            dlg.accept()
        def on_no_photo():
            result[0] = "batch"
            dlg.accept()
        btn_direct.clicked.connect(on_direct)
        btn_no_photo.clicked.connect(on_no_photo)
        btn_cancel.clicked.connect(dlg.reject)
        if dlg.exec_() != QDialog.Accepted or result[0] is None:
            return
        mode = result[0]
        file_path, _ = QFileDialog.getOpenFileName(self, "选择人员导入文件", "", "Excel文件 (*.xlsx *.xls)")
        if not file_path:
            return
        temp_photo_dir = None
        if mode == "batch":
            temp_photo_dir = os.path.join(BASE_DIR, f"头像批量导入_{datetime.now().strftime('%H%M%S')}")
            os.makedirs(temp_photo_dir, exist_ok=True)
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp_photo_dir))
            QMessageBox.information(self, "提示", f"请将照片放入打开的文件夹中。\n照片命名格式：img (1).jpg，img (2).jpg ...\n\n点击确定后开始导入并匹配头像。")
        cnt = self.pm.import_from_excel(file_path, mode, temp_photo_dir if mode == "batch" else None)
        if cnt > 0:
            QMessageBox.information(self, "完成", f"成功导入 {cnt} 条人员记录。")
            if mode == "batch" and temp_photo_dir:
                self.match_avatars(temp_photo_dir)
                try:
                    shutil.rmtree(temp_photo_dir)
                except:
                    pass
            self.refresh_list()
        else:
            QMessageBox.warning(self, "失败", "未导入任何人员，请检查文件格式。")

    def match_avatars(self, photo_dir):
        matched = 0
        for person in self.pm.persons:
            seq = int(person.id)
            for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                for prefix in ['img ', '人员img ']:
                    test_name = f"{prefix}({seq}){ext}"
                    test_path = os.path.join(photo_dir, test_name)
                    if os.path.exists(test_path):
                        new_filename = f"{person.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                        dest = os.path.join(PHOTO_DIR, new_filename)
                        shutil.copy2(test_path, dest)
                        if person.photo:
                            old_path = os.path.join(PHOTO_DIR, person.photo)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        person.photo = new_filename
                        matched += 1
                        break
                if person.photo:
                    break
        self.pm.save()
        QMessageBox.information(self, "匹配完成", f"共匹配 {matched} 个头像。")
        self.refresh_list()

class PersonEditDialog(QDialog):
    def __init__(self, person_manager, person=None, parent=None):
        super().__init__(parent)
        self.pm = person_manager
        self.person = person
        self.setWindowTitle("编辑人员" if person else "新增人员")
        self.setMinimumWidth(400)
        self.photo_path = None
        self.init_ui()
        if person:
            self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.nickname_edit = QLineEdit()
        layout.addWidget(QLabel("昵称(*):"))
        layout.addWidget(self.nickname_edit)
        self.realname_edit = QLineEdit()
        layout.addWidget(QLabel("真实姓名:"))
        layout.addWidget(self.realname_edit)
        self.avatar_label = QLabel()
        self.avatar_label.setFixedSize(100, 100)
        self.avatar_label.setStyleSheet("border: 1px solid gray;")
        self.avatar_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(QLabel("头像:"))
        layout.addWidget(self.avatar_label)
        self.btn_select_photo = QPushButton("选择照片")
        self.btn_select_photo.clicked.connect(self.select_photo)
        layout.addWidget(self.btn_select_photo)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def load_data(self):
        self.nickname_edit.setText(self.person.nickname)
        self.realname_edit.setText(self.person.realname)
        if self.person.photo:
            photo_path = os.path.join(PHOTO_DIR, self.person.photo)
            if os.path.exists(photo_path):
                pix = QPixmap(photo_path).scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.avatar_label.setPixmap(pix)

    def select_photo(self):
        settings = load_settings()
        last_dir = settings.get("last_photo_dir", os.path.expanduser("~"))
        file_path, _ = QFileDialog.getOpenFileName(self, "选择照片", last_dir, "图片文件 (*.png *.jpg *.jpeg *.bmp)")
        if file_path:
            settings["last_photo_dir"] = os.path.dirname(file_path)
            save_settings(settings)
            crop_dlg = CropDialog(file_path, self)
            if crop_dlg.exec_():
                temp_path = os.path.join(PHOTO_DIR, "_temp_crop.png")
                crop_dlg.save_cropped(temp_path)
                self.photo_path = temp_path
                pix = QPixmap(temp_path).scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.avatar_label.setPixmap(pix)

    def accept(self):
        nickname = self.nickname_edit.text().strip()
        if not nickname:
            QMessageBox.warning(self, "警告", "昵称不能为空")
            return
        realname = self.realname_edit.text().strip()
        if self.person:
            self.pm.update_person(self.person.id, nickname, realname, self.photo_path)
        else:
            self.pm.add_person(nickname, realname, self.photo_path)
        if self.photo_path and self.photo_path.endswith("_temp_crop.png") and os.path.exists(self.photo_path):
            os.remove(self.photo_path)
        super().accept()

# ==================== 统计对话框 ====================
class StatsDialog(QDialog):
    def __init__(self, person_manager, parent=None):
        super().__init__(parent)
        self.pm = person_manager
        self.setWindowTitle("数据统计")
        self.setMinimumSize(800, 600)
        self.init_ui()
        self.load_stats()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.tab_widget = QTabWidget()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(5)
        self.table_widget.setHorizontalHeaderLabels(["ID", "昵称", "真实姓名", "中选次数", "中选概率"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_widget.setSortingEnabled(True)
        self.tab_widget.addTab(self.table_widget, "详细数据")
        self.figure = Figure(figsize=(9, 6), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        self.tab_widget.addTab(self.canvas, "柱状图")
        layout.addWidget(self.tab_widget)
        btn_layout = QHBoxLayout()
        self.btn_clear = QPushButton("清空记录")
        self.btn_clear.clicked.connect(self.clear_stats)
        btn_layout.addWidget(self.btn_clear)
        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(self.btn_close)
        layout.addLayout(btn_layout)

    def clear_stats(self):
        if QMessageBox.question(self, "确认清空", "清空后将丢失所有历史抽选记录，是否继续？") == QMessageBox.Yes:
            with open(STATS_FILE, 'w', encoding='utf-8') as f:
                json.dump([], f)
            self.load_stats()

    def load_stats(self):
        stats = load_stats()
        total_picks = 0
        count_dict = {}
        for entry in stats:
            chosen_list = entry.get("chosen_ids", [])
            total_picks += len(chosen_list)
            for chosen_id in chosen_list:
                count_dict[chosen_id] = count_dict.get(chosen_id, 0) + 1
        self.table_widget.setRowCount(len(self.pm.persons))
        for row, person in enumerate(self.pm.persons):
            cnt = count_dict.get(person.id, 0)
            prob = (cnt / total_picks * 100) if total_picks > 0 else 0.0
            id_item = QTableWidgetItem(person.id)
            id_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 0, id_item)
            nick_item = QTableWidgetItem(person.nickname)
            nick_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 1, nick_item)
            real_item = QTableWidgetItem(person.realname)
            real_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 2, real_item)
            cnt_item = QTableWidgetItem(str(cnt))
            cnt_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 3, cnt_item)
            prob_item = QTableWidgetItem(f"{prob:.2f}%")
            prob_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 4, prob_item)
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        names = [p.nickname for p in self.pm.persons]
        probs = [count_dict.get(p.id, 0) / total_picks * 100 if total_picks > 0 else 0 for p in self.pm.persons]
        ax.bar(names, probs, color='steelblue')
        ax.set_ylabel('中选概率 (%)')
        ax.set_xlabel('人员')
        ax.set_title('中选概率分布')
        ax.tick_params(axis='x', rotation=45, labelsize=9)
        self.figure.subplots_adjust(bottom=0.25)
        for i, v in enumerate(probs):
            if v > 0:
                offset = max(0.5, v * 0.03)
                ax.text(i, v + offset, f"{v:.1f}%", ha='center', va='bottom', fontsize=8)
        self.figure.tight_layout()
        self.canvas.draw()

# ==================== 结果弹窗（非模态，绝对居中）====================
class ResultDialog(QDialog):
    def __init__(self, person, main_window, auto_close_seconds=0):
        super().__init__(main_window)
        self.person = person
        self.main_window = main_window
        self.auto_close_seconds = auto_close_seconds

        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setModal(False)

        parent_geo = self.main_window.geometry()
        base_w = parent_geo.width()
        base_h = parent_geo.height()
        w = max(400, min(600, int(base_w * 0.33)))
        h = max(500, min(600, int(base_h * 0.5)))
        self.resize(w, h)

        self.init_ui()

        if auto_close_seconds > 0:
            QTimer.singleShot(int(auto_close_seconds * 1000), self.accept)

    def center(self):
        parent_rect = self.parent().geometry()
        x = parent_rect.center().x() - self.width() // 2
        y = parent_rect.center().y() - self.height() // 2
        self.move(x, y)

    def showEvent(self, event):
        self.center()
        super().showEvent(event)

    def resizeEvent(self, event):
        self.center()
        super().resizeEvent(event)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 30, 20, 30)
        card = QFrame()
        card.setObjectName("resultCard")
        card.setStyleSheet("""
            #resultCard {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #FFDEE9, stop:1 #B5FFFC);
                border-radius: 20px;
                border: 2px solid white;
            }
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(25)
        title_label = QLabel("🏆 抽取结果 🏆")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #c2185b;")
        card_layout.addWidget(title_label)
        avatar_path = os.path.join(PHOTO_DIR, self.person.photo) if self.person.photo else ""
        if os.path.exists(avatar_path):
            pix = QPixmap(avatar_path).scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        else:
            pix = QPixmap(180, 180)
            pix.fill(Qt.lightGray)
        avatar_label = QLabel()
        avatar_label.setPixmap(pix)
        avatar_label.setAlignment(Qt.AlignCenter)
        avatar_label.setStyleSheet("border-radius: 90px; border: 3px solid white;")
        card_layout.addWidget(avatar_label)
        settings = load_settings()
        show_nickname = settings.get("draw_show_nickname", True)
        show_realname = settings.get("draw_show_realname", True)
        if show_nickname and show_realname and self.person.realname:
            name = f"{self.person.nickname} ({self.person.realname})"
        elif show_realname and self.person.realname:
            name = self.person.realname
        else:
            name = self.person.nickname
        name_label = QLabel(name)
        name_label.setAlignment(Qt.AlignCenter)
        name_label.setStyleSheet("font-size: 28px; font-weight: bold; color: #d32f2f;")
        card_layout.addWidget(name_label)
        btn_close = QPushButton("✕ 关闭")
        btn_close.setStyleSheet("""
            QPushButton {
                background: #ff4081;
                color: white;
                font-size: 16px;
                padding: 8px;
                border-radius: 20px;
                border: none;
                min-width: 100px;
            }
            QPushButton:hover {
                background: #c2185b;
            }
        """)
        btn_close.clicked.connect(self.accept)
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(btn_close)
        button_layout.addStretch()
        card_layout.addLayout(button_layout)
        main_layout.addWidget(card)

# ==================== 卡片组件 ====================
class PersonCard(QFrame):
    def __init__(self, person, parent=None):
        super().__init__(parent)
        self.person = person
        self.setFrameShape(QFrame.Box)
        self.setFixedSize(120, 150)
        self.setStyleSheet("PersonCard { border: 1px solid #ccc; border-radius: 8px; background: white; }")
        layout = QVBoxLayout(self)
        self.avatar_label = QLabel()
        self.avatar_label.setFixedSize(80, 80)
        self.avatar_label.setAlignment(Qt.AlignCenter)
        self.load_avatar()
        layout.addWidget(self.avatar_label, alignment=Qt.AlignCenter)
        self.name_label = QLabel()
        self.name_label.setAlignment(Qt.AlignCenter)
        self.name_label.setWordWrap(True)
        self.update_name_display()
        layout.addWidget(self.name_label)

    def update_name_display(self):
        settings = load_settings()
        show_nick = settings.get("card_show_nickname", True)
        show_real = settings.get("card_show_realname", False)
        if show_nick and show_real and self.person.realname:
            text = f"{self.person.nickname}\n({self.person.realname})"
        elif show_real and self.person.realname:
            text = self.person.realname
        else:
            text = self.person.nickname
        self.name_label.setText(text)

    def load_avatar(self):
        if self.person.photo:
            photo_path = os.path.join(PHOTO_DIR, self.person.photo)
            if os.path.exists(photo_path):
                pix = QPixmap(photo_path).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.avatar_label.setPixmap(pix)
                return
        img = Image.new('RGB', (80, 80), color=(100,100,200))
        draw = ImageDraw.Draw(img)
        char = self.person.nickname[0] if self.person.nickname else "?"
        try:
            font = ImageFont.truetype("arial.ttf", 40)
        except:
            font = ImageFont.load_default()
        bbox = draw.textbbox((0,0), char, font=font)
        tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
        draw.text(((80-tw)//2, (80-th)//2), char, fill=(255,255,255), font=font)
        img = img.convert("RGBA")
        data = img.tobytes("raw", "RGBA")
        qim = QImage(data, 80, 80, QImage.Format_RGBA8888)
        self.avatar_label.setPixmap(QPixmap.fromImage(qim))

    def set_highlight(self, enabled):
        if enabled:
            self.setStyleSheet("PersonCard { border: 4px solid #ff5722; border-radius: 8px; background: #fff5f0; }")
        else:
            self.setStyleSheet("PersonCard { border: 1px solid #ccc; border-radius: 8px; background: white; }")

# ==================== 显示设置对话框 ====================
class SettingsDialog(QDialog):
    def __init__(self, main_window, person_manager, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.pm = person_manager
        self.setWindowTitle("显示设置")
        self.setMinimumWidth(720)
        layout = QVBoxLayout(self)

        group_card = QGroupBox("人员名片显示")
        card_layout = QHBoxLayout(group_card)
        self.cb_card_nick = QCheckBox("显示昵称")
        self.cb_card_real = QCheckBox("显示真实姓名")
        card_layout.addWidget(self.cb_card_nick)
        card_layout.addWidget(self.cb_card_real)
        layout.addWidget(group_card)

        group_last = QGroupBox("上次结果显示")
        last_layout = QHBoxLayout(group_last)
        self.cb_last_nick = QCheckBox("显示昵称")
        self.cb_last_real = QCheckBox("显示真实姓名")
        last_layout.addWidget(self.cb_last_nick)
        last_layout.addWidget(self.cb_last_real)
        layout.addWidget(group_last)

        group_draw = QGroupBox("抽中结果显示")
        draw_layout = QHBoxLayout(group_draw)
        self.cb_draw_nick = QCheckBox("显示昵称")
        self.cb_draw_real = QCheckBox("显示真实姓名")
        draw_layout.addWidget(self.cb_draw_nick)
        draw_layout.addWidget(self.cb_draw_real)
        layout.addWidget(group_draw)

        group_single = QGroupBox("单次抽取弹窗自动关闭")
        single_layout = QHBoxLayout(group_single)
        self.cb_auto_close_single = QCheckBox("启用自动关闭")
        self.spin_single_close = QDoubleSpinBox()
        self.spin_single_close.setRange(0.5, 30)
        self.spin_single_close.setSingleStep(0.5)
        self.spin_single_close.setValue(2.0)
        self.spin_single_close.setSuffix(" 秒")
        self.spin_single_close.setEnabled(False)
        self.cb_auto_close_single.toggled.connect(self.spin_single_close.setEnabled)
        single_layout.addWidget(self.cb_auto_close_single)
        single_layout.addWidget(self.spin_single_close)
        single_layout.addStretch()
        layout.addWidget(group_single)

        group_multi = QGroupBox("多次抽取弹窗自动关闭")
        multi_layout = QHBoxLayout(group_multi)
        self.cb_auto_close_multi = QCheckBox("启用自动关闭")
        self.spin_multi_close = QDoubleSpinBox()
        self.spin_multi_close.setRange(0.1, 59.0)
        self.spin_multi_close.setSingleStep(0.1)
        self.spin_multi_close.setValue(0.5)
        self.spin_multi_close.setSuffix(" 秒")
        self.spin_multi_close.setEnabled(True)   # 默认启用，所以时间框可用
        self.cb_auto_close_multi.toggled.connect(self.spin_multi_close.setEnabled)
        multi_layout.addWidget(self.cb_auto_close_multi)
        multi_layout.addWidget(self.spin_multi_close)
        multi_layout.addStretch()
        layout.addWidget(group_multi)

        group_anim_speed = QGroupBox("动画速度（毫秒/步）")
        anim_speed_layout = QHBoxLayout(group_anim_speed)
        self.spin_speed_front = QSpinBox()
        self.spin_speed_front.setRange(10, 500)
        self.spin_speed_front.setValue(40)
        self.spin_speed_front.setSuffix(" ms")
        self.spin_speed_mid = QSpinBox()
        self.spin_speed_mid.setRange(10, 500)
        self.spin_speed_mid.setValue(80)
        self.spin_speed_mid.setSuffix(" ms")
        self.spin_speed_back = QSpinBox()
        self.spin_speed_back.setRange(10, 500)
        self.spin_speed_back.setValue(120)
        self.spin_speed_back.setSuffix(" ms")
        anim_speed_layout.addWidget(QLabel("前段:"))
        anim_speed_layout.addWidget(self.spin_speed_front)
        anim_speed_layout.addWidget(QLabel("中段:"))
        anim_speed_layout.addWidget(self.spin_speed_mid)
        anim_speed_layout.addWidget(QLabel("后段:"))
        anim_speed_layout.addWidget(self.spin_speed_back)
        anim_speed_layout.addStretch()
        layout.addWidget(group_anim_speed)

        group_timing = QGroupBox("抽取结果显示时机（右侧面板）")
        timing_layout = QHBoxLayout(group_timing)
        self.radio_immediate = QRadioButton("即时显示")
        self.radio_delayed = QRadioButton("延后显示")
        timing_layout.addWidget(self.radio_immediate)
        timing_layout.addWidget(self.radio_delayed)
        layout.addWidget(group_timing)

        group_show = QGroupBox("右侧面板组件显示")
        show_layout = QVBoxLayout(group_show)
        self.cb_show_history = QCheckBox("显示多组结果")
        self.cb_show_process = QCheckBox("显示抽取过程记录")
        show_layout.addWidget(self.cb_show_history)
        show_layout.addWidget(self.cb_show_process)
        layout.addWidget(group_show)

        self.cb_title_suffix = QCheckBox("显示标题后缀「CopyRigth © XAF 2026.5」")
        self.cb_log_panel = QCheckBox("显示抽取记录面板（右侧整体）")
        layout.addWidget(self.cb_title_suffix)
        layout.addWidget(self.cb_log_panel)

        button_layout = QHBoxLayout()
        self.btn_reset_size = QPushButton("重置窗口大小")
        self.btn_reset_size.clicked.connect(self.reset_window_size)
        self.btn_reset_config = QPushButton("重置配置")
        self.btn_reset_config.clicked.connect(self.reset_config)
        button_layout.addWidget(self.btn_reset_size)
        button_layout.addWidget(self.btn_reset_config)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.load_settings()
        self.setFixedHeight(self.sizeHint().height())

    def reset_config(self):
        reply = QMessageBox.question(self, "重置配置", 
            "重置所有配置将恢复所有显示选项、自动关闭时间、动画速度等为默认值。\n窗口大小和分割线位置不会被重置。\n人员信息和抽选记录不受影响。\n确定要重置吗？",
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            current_geo = self.main_window.saveGeometry().toHex().data().decode()
            current_main_sizes = self.main_window.main_splitter.sizes()
            current_right_sizes = self.main_window.right_splitter.sizes()
            settings = load_settings()
            preserved = {
                "window_geometry": current_geo,
                "main_splitter_sizes": current_main_sizes,
                "right_splitter_sizes": current_right_sizes,
                "last_photo_dir": settings.get("last_photo_dir", os.path.expanduser("~"))
            }
            default_settings = {
                "last_photo_dir": preserved["last_photo_dir"],
                "card_show_nickname": True,
                "card_show_realname": False,
                "last_show_nickname": True,
                "last_show_realname": False,
                "draw_show_nickname": True,
                "draw_show_realname": False,
                "log_panel_visible": True,
                "draw_k": 1,
                "draw_repeat": 1,
                "show_title_copyright": True,
                "auto_close_single_enabled": False,
                "auto_close_single_seconds": 2.0,
                "auto_close_multi_enabled": True,       # 多次抽取默认启用
                "auto_close_multi_seconds": 0.5,
                "anim_speed_front": 40,
                "anim_speed_mid": 80,
                "anim_speed_back": 120,
                "result_display_immediate": False,
                "show_history_list": True,
                "show_process_log": True,
                "window_geometry": preserved["window_geometry"],
                "main_splitter_sizes": preserved["main_splitter_sizes"],
                "right_splitter_sizes": preserved["right_splitter_sizes"]
            }
            save_settings(default_settings)
            self.load_settings()
            self.main_window.update_window_title()
            self.main_window.update_right_panel_visibility()
            self.main_window.toggle_log_action.setChecked(default_settings["log_panel_visible"])
            self.main_window.toggle_process_panel(default_settings["log_panel_visible"])
            self.main_window.refresh_cards()
            self.main_window.update_last_result_display()
            QMessageBox.information(self, "重置完成", "配置已重置，部分更改需要关闭设置对话框后生效。")

    def load_settings(self):
        settings = load_settings()
        self.cb_card_nick.setChecked(settings.get("card_show_nickname", True))
        self.cb_card_real.setChecked(settings.get("card_show_realname", False))
        self.cb_last_nick.setChecked(settings.get("last_show_nickname", True))
        self.cb_last_real.setChecked(settings.get("last_show_realname", False))
        self.cb_draw_nick.setChecked(settings.get("draw_show_nickname", True))
        self.cb_draw_real.setChecked(settings.get("draw_show_realname", False))
        self.cb_title_suffix.setChecked(settings.get("show_title_copyright", True))
        self.cb_log_panel.setChecked(settings.get("log_panel_visible", True))
        self.cb_auto_close_single.setChecked(settings.get("auto_close_single_enabled", False))
        self.spin_single_close.setValue(settings.get("auto_close_single_seconds", 2.0))
        self.cb_auto_close_multi.setChecked(settings.get("auto_close_multi_enabled", True))
        self.spin_multi_close.setValue(settings.get("auto_close_multi_seconds", 0.5))
        self.spin_multi_close.setEnabled(self.cb_auto_close_multi.isChecked())
        self.spin_speed_front.setValue(settings.get("anim_speed_front", 40))
        self.spin_speed_mid.setValue(settings.get("anim_speed_mid", 80))
        self.spin_speed_back.setValue(settings.get("anim_speed_back", 120))
        self.radio_immediate.setChecked(settings.get("result_display_immediate", False))
        self.radio_delayed.setChecked(not settings.get("result_display_immediate", False))
        self.cb_show_history.setChecked(settings.get("show_history_list", True))
        self.cb_show_process.setChecked(settings.get("show_process_log", True))

    def reset_window_size(self):
        self.main_window.resize(1200, 800)
        self.main_window.center_window()

    def accept(self):
        settings = load_settings()
        settings["card_show_nickname"] = self.cb_card_nick.isChecked()
        settings["card_show_realname"] = self.cb_card_real.isChecked()
        settings["last_show_nickname"] = self.cb_last_nick.isChecked()
        settings["last_show_realname"] = self.cb_last_real.isChecked()
        settings["draw_show_nickname"] = self.cb_draw_nick.isChecked()
        settings["draw_show_realname"] = self.cb_draw_real.isChecked()
        settings["show_title_copyright"] = self.cb_title_suffix.isChecked()
        settings["log_panel_visible"] = self.cb_log_panel.isChecked()
        settings["auto_close_single_enabled"] = self.cb_auto_close_single.isChecked()
        settings["auto_close_single_seconds"] = self.spin_single_close.value()
        settings["auto_close_multi_enabled"] = self.cb_auto_close_multi.isChecked()
        settings["auto_close_multi_seconds"] = self.spin_multi_close.value()
        settings["anim_speed_front"] = self.spin_speed_front.value()
        settings["anim_speed_mid"] = self.spin_speed_mid.value()
        settings["anim_speed_back"] = self.spin_speed_back.value()
        settings["result_display_immediate"] = self.radio_immediate.isChecked()
        settings["show_history_list"] = self.cb_show_history.isChecked()
        settings["show_process_log"] = self.cb_show_process.isChecked()
        save_settings(settings)

        self.main_window.refresh_cards()
        self.main_window.update_last_result_display()
        self.main_window.update_window_title()
        self.main_window.toggle_log_action.setChecked(settings["log_panel_visible"])
        self.main_window.toggle_process_panel(settings["log_panel_visible"])
        self.main_window.update_right_panel_visibility()
        super().accept()

# ==================== 关于对话框 ====================
class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("关于")
        self.setFixedSize(550, 480)
        layout = QVBoxLayout(self)
        text = QLabel("""
        <center>
        <h2>随机抽选工具</h2>
        <p><b>版本：</b>1.3.22</p>
        <p><b>作者：</b>XAF</p>
        <p><b>制作日期：</b>2026年5月4日</p>
        <hr>
        <h3>抽取机制（双随机）</h3>
        <p>本工具采用基于微秒级时间种子的安全随机算法：</p>
        <ul style="display: inline-block; text-align: left;">
        <li><b>映射随机：</b>使用 <code>secrets</code> 模块对参与人员ID进行随机排序，生成序号→人员映射，每次抽取均重新打乱。</li>
        <li><b>抽取随机：</b>以当前时间微秒值（1~999999）作为随机源，按参与人数平分区间，余数丢弃。微秒值落在谁的区域即抽中谁。</li>
        <li>多人抽取时自动去除已中选者（不放回），剩余人员重新映射序号，确保每个抽取步骤独立公正。</li>
        </ul>
        <i>本工具完全离线运行，日志记录每一步。</i>
        </center>
        """)
        text.setWordWrap(True)
        text.setOpenExternalLinks(True)
        layout.addWidget(text)
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self, person_manager, picker, logger):
        super().__init__()
        self.pm = person_manager
        self.picker = picker
        self.logger = logger
        self.cards = {}
        self.animation_timer = QTimer()
        self.animation_timer.timeout.connect(self.animate_step)
        self.animation_index = 0
        self.animation_sequence = []
        self.animation_target_id = None
        self.animation_speed = 50
        self.is_animating = False
        self.result_callback = None
        self.multi_results = []
        self.current_draw_index = 0
        self.repeat_times = 1
        self.all_draw_results = []
        self.defer_logs = False
        self.pending_logs = []
        self.pending_draw_result = None
        self.current_result_dialog = None
        self.setWindowIcon(QIcon(ICON_PATH) if os.path.exists(ICON_PATH) else QIcon())
        self.init_ui()
        self.load_window_state()
        self.refresh_cards()
        self.create_tray()
        self.start_local_server()
        self.timer_clock = QTimer()
        self.timer_clock.timeout.connect(self.update_clock)
        self.timer_clock.start(100)
        settings = load_settings()
        self.toggle_log_action.setChecked(settings.get("log_panel_visible", True))
        self.toggle_process_panel(settings.get("log_panel_visible", True))
        self.spin_k.setValue(settings.get("draw_k", 1))
        self.spin_repeat.setValue(settings.get("draw_repeat", 1))
        self.update_window_title()
        self.update_right_panel_visibility()

    def update_window_title(self):
        settings = load_settings()
        base_title = "随机抽取工具"
        if settings.get("show_title_copyright", True):
            self.setWindowTitle(f"{base_title}  CopyRigth © XAF 2026.5")
        else:
            self.setWindowTitle(base_title)

    def update_right_panel_visibility(self):
        settings = load_settings()
        self.history_group.setVisible(settings.get("show_history_list", True))
        self.process_group.setVisible(settings.get("show_process_log", True))

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)

        toolbar = self.addToolBar("工具栏")
        toolbar.addAction("人员维护", self.open_person_manage)
        toolbar.addAction("数据统计", self.open_stats)
        toolbar.addAction("显示设置", self.open_settings)
        self.toggle_log_action = toolbar.addAction("抽取记录")
        self.toggle_log_action.setCheckable(True)
        self.toggle_log_action.toggled.connect(self.toggle_process_panel)
        toolbar.addAction("关于", self.open_about)

        self.scroll_area = QScrollArea()
        self.card_container = QWidget()
        self.card_layout = QGridLayout(self.card_container)
        self.scroll_area.setWidget(self.card_container)
        self.scroll_area.setWidgetResizable(True)
        left_layout.addWidget(self.scroll_area)

        control_group = QGroupBox("抽取选择")
        control_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 14px; }")
        control_layout = QHBoxLayout(control_group)
        control_layout.setSpacing(20)
        k_label = QLabel("抽取人数:")
        k_label.setFixedWidth(70)
        control_layout.addWidget(k_label)
        self.spin_k = QSpinBox()
        self.spin_k.setMinimum(1)
        self.spin_k.setValue(1)
        self.spin_k.setFixedWidth(70)
        control_layout.addWidget(self.spin_k)
        repeat_label = QLabel("抽取次数:")
        repeat_label.setFixedWidth(70)
        control_layout.addWidget(repeat_label)
        self.spin_repeat = QSpinBox()
        self.spin_repeat.setMinimum(1)
        self.spin_repeat.setMaximum(100)
        self.spin_repeat.setValue(1)
        self.spin_repeat.setFixedWidth(70)
        control_layout.addWidget(self.spin_repeat)
        self.btn_draw = QPushButton("开始抽取")
        self.btn_draw.setFixedHeight(50)
        self.btn_draw.setStyleSheet("""
            QPushButton {
                font-size: 18px;
                font-weight: bold;
                background: #4CAF50;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background: #45a049;
            }
        """)
        self.btn_draw.clicked.connect(self.start_draw)
        control_layout.addWidget(self.btn_draw)
        self.last_result_label = QLabel("上次结果: 无")
        self.last_result_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #d32f2f; background: #fff3e0; padding: 5px 10px; border-radius: 8px;")
        self.last_result_label.setAlignment(Qt.AlignCenter)
        control_layout.addWidget(self.last_result_label)
        left_layout.addWidget(control_group)

        self.right_panel = QWidget()
        self.right_panel.setMaximumWidth(550)
        right_layout = QVBoxLayout(self.right_panel)

        top_group = QGroupBox("实时信息")
        top_layout = QVBoxLayout(top_group)
        self.clock_label = QLabel()
        self.clock_label.setFont(QFont("Consolas", 20))
        self.clock_label.setAlignment(Qt.AlignCenter)
        self.clock_label.setStyleSheet("color: #2c3e50; font-weight: bold;")
        top_layout.addWidget(self.clock_label)
        self.current_result_label = QLabel("暂无")
        self.current_result_label.setWordWrap(True)
        self.current_result_label.setStyleSheet("font-size: 14px; background: #fff3e0; padding: 8px; border-radius: 8px;")
        self.current_result_label.setAlignment(Qt.AlignCenter)
        top_layout.addWidget(QLabel("📌 本次抽取结果:"))
        top_layout.addWidget(self.current_result_label)
        right_layout.addWidget(top_group)

        self.history_group = QGroupBox("多组结果")
        history_layout = QVBoxLayout(self.history_group)
        self.history_grid = QWidget()
        self.history_grid_layout = QGridLayout(self.history_grid)
        self.history_grid_layout.setSpacing(5)
        history_layout.addWidget(self.history_grid)
        right_layout.addWidget(self.history_group)

        self.process_group = QGroupBox("抽取过程")
        process_layout = QVBoxLayout(self.process_group)
        self.process_log = QTextEdit()
        self.process_log.setReadOnly(True)
        self.process_log.setFont(QFont("Consolas", 9))
        process_layout.addWidget(self.process_log)
        clear_btn = QPushButton("清空记录")
        clear_btn.clicked.connect(self.clear_process_log)
        process_layout.addWidget(clear_btn)
        right_layout.addWidget(self.process_group)

        self.right_splitter = QSplitter(Qt.Vertical)
        self.right_splitter.addWidget(self.history_group)
        self.right_splitter.addWidget(self.process_group)
        self.right_splitter.setSizes([300, 400])
        self.right_splitter.splitterMoved.connect(self.save_right_splitter_state)
        right_layout.removeWidget(self.history_group)
        right_layout.removeWidget(self.process_group)
        right_layout.addWidget(self.right_splitter)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.addWidget(left_widget)
        self.main_splitter.addWidget(self.right_panel)
        self.main_splitter.setSizes([800, 400])
        self.main_splitter.splitterMoved.connect(self.save_main_splitter_state)
        main_layout.addWidget(self.main_splitter)

        self.picker.ui_callback = self.append_process_log

    def update_clock(self):
        now = datetime.now()
        self.clock_label.setText(now.strftime("%H:%M:%S.%f")[:-3])

    def toggle_process_panel(self, checked):
        self.right_panel.setVisible(checked)
        settings = load_settings()
        settings["log_panel_visible"] = checked
        save_settings(settings)

    def clear_process_log(self):
        self.process_log.clear()

    def append_process_log(self, text, timestamp=True, deferrable=True):
        if timestamp:
            time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_line = f"[{time_str}] {text}"
        else:
            log_line = text
        if "→ 抽中:" in log_line:
            log_line += "\n"
        if self.defer_logs and deferrable:
            self.pending_logs.append(log_line)
        else:
            self.process_log.append(log_line)
            self.logger.log_raw(log_line)

    def flush_pending_logs(self):
        if self.pending_logs:
            self.process_log.append("\n".join(self.pending_logs))
            for line in self.pending_logs:
                self.logger.log_raw(line)
            self.pending_logs.clear()

    def refresh_cards(self):
        for i in reversed(range(self.card_layout.count())):
            widget = self.card_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        self.cards.clear()
        participants = [p for p in self.pm.persons if p.participate]
        col = 0
        row = 0
        for person in participants:
            card = PersonCard(person)
            self.card_layout.addWidget(card, row, col)
            self.cards[person.id] = card
            col += 1
            if col >= 5:
                col = 0
                row += 1
        self.spin_k.setMaximum(max(1, len(participants)))

    def open_person_manage(self):
        dlg = PersonManageDialog(self.pm, self)
        dlg.exec_()
        self.refresh_cards()

    def open_stats(self):
        dlg = StatsDialog(self.pm, self)
        dlg.exec_()

    def open_settings(self):
        dlg = SettingsDialog(self, self.pm, self)
        dlg.exec_()
        self.update_last_result_display()
        self.update_history_display()

    def open_about(self):
        dlg = AboutDialog(self)
        dlg.exec_()

    def update_last_result_display(self):
        if hasattr(self, 'last_draw_person') and self.last_draw_person:
            settings = load_settings()
            show_nick = settings.get("last_show_nickname", True)
            show_real = settings.get("last_show_realname", False)
            if show_nick and show_real and self.last_draw_person.realname:
                name = f"{self.last_draw_person.nickname} ({self.last_draw_person.realname})"
            elif show_real and self.last_draw_person.realname:
                name = self.last_draw_person.realname
            else:
                name = self.last_draw_person.nickname
            self.last_result_label.setText(f"上次结果: {name}")

    def start_draw(self):
        if self.is_animating:
            return
        k = self.spin_k.value()
        repeat = self.spin_repeat.value()
        participant_ids = self.pm.get_participants()
        if len(participant_ids) < k:
            QMessageBox.warning(self, "警告", f"参与抽取人数不足，当前{len(participant_ids)}人，需要{k}人。")
            return
        settings = load_settings()
        settings["draw_k"] = k
        settings["draw_repeat"] = repeat
        save_settings(settings)

        self.btn_draw.setEnabled(False)
        self.repeat_times = repeat
        self.current_draw_index = 0
        self.all_draw_results = []
        self.pending_draw_result = None
        self.update_history_display()
        self.current_result_label.setText("暂无")
        self.do_next_draw()

    def do_next_draw(self):
        if self.current_draw_index >= self.repeat_times:
            self.btn_draw.setEnabled(True)
            return
        self.flush_pending_logs()
        self.append_process_log(f"\n========== 第 {self.current_draw_index+1}/{self.repeat_times} 次抽取 ==========", timestamp=True, deferrable=False)
        participant_ids = self.pm.get_participants()
        k = self.spin_k.value()
        settings = load_settings()
        immediate = settings.get("result_display_immediate", False)
        self.defer_logs = not immediate
        if not self.defer_logs:
            self.pending_logs.clear()
        else:
            self.pending_logs = []

        chosen_ids = self.picker.pick(participant_ids, k, with_replacement=False)
        record_draw_result(chosen_ids)
        self.logger.log_draw({"final_chosen": chosen_ids, "draw_index": self.current_draw_index+1})

        if immediate:
            self.all_draw_results.append(chosen_ids)
            self.update_history_display()
        else:
            self.pending_draw_result = chosen_ids

        if len(chosen_ids) == 1:
            self.start_animation(chosen_ids[0], self.on_single_draw_finished)
        else:
            self.multi_results = chosen_ids
            self.current_multi_index = 0
            self.show_next_multi_result()

    def format_person_name(self, person, settings):
        show_nick = settings.get("draw_show_nickname", True)
        show_real = settings.get("draw_show_realname", False)
        if show_nick and show_real and person.realname:
            return f"{person.nickname} ({person.realname})"
        elif show_real and person.realname:
            return person.realname
        else:
            return person.nickname

    def update_history_display(self):
        for i in reversed(range(self.history_grid_layout.count())):
            widget = self.history_grid_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        if not self.all_draw_results:
            empty_label = QLabel("无")
            empty_label.setAlignment(Qt.AlignCenter)
            self.history_grid_layout.addWidget(empty_label, 0, 0)
            return
        settings = load_settings()
        row = 0
        col = 0
        for idx, ids in enumerate(self.all_draw_results, 1):
            names = []
            for pid in ids:
                person = self.pm.get_person_by_id(pid)
                if person:
                    names.append(self.format_person_name(person, settings))
            text = f"{idx}. {', '.join(names)}"
            label = QLabel(text)
            label.setWordWrap(True)
            label.setStyleSheet("background: #f5f5f5; padding: 4px; border-radius: 4px; margin: 2px;")
            label.setAlignment(Qt.AlignCenter)
            self.history_grid_layout.addWidget(label, row, col)
            col += 1
            if col >= 5:
                col = 0
                row += 1

    def close_current_result_dialog(self):
        if self.current_result_dialog is not None:
            try:
                self.current_result_dialog.accept()
                self.current_result_dialog.deleteLater()
            except:
                pass
            self.current_result_dialog = None

    def on_single_draw_finished(self):
        settings = load_settings()
        immediate = settings.get("result_display_immediate", False)
        if not immediate and self.pending_draw_result is not None:
            self.all_draw_results.append(self.pending_draw_result)
            self.update_history_display()
            self.pending_draw_result = None
            self.flush_pending_logs()
            self.defer_logs = False

        pid = self.all_draw_results[self.current_draw_index][0]
        person = self.pm.get_person_by_id(pid)
        self.last_draw_person = person
        self.update_last_result_display()
        if not immediate:
            self.current_result_label.setText(f"第{self.current_draw_index+1}次: {self.format_person_name(person, settings)}")
        else:
            if not settings.get("result_display_immediate", False):
                self.current_result_label.setText(f"第{self.current_draw_index+1}次: {self.format_person_name(person, settings)}")

        self.close_current_result_dialog()
        auto_close = settings.get("auto_close_single_seconds", 2.0) if settings.get("auto_close_single_enabled", False) else 0
        self.current_result_dialog = ResultDialog(person, self, auto_close)
        self.current_result_dialog.show()
        self.current_draw_index += 1
        QTimer.singleShot(500, self.do_next_draw)

    def show_next_multi_result(self):
        if self.current_multi_index < len(self.multi_results):
            pid = self.multi_results[self.current_multi_index]
            self.start_animation(pid, self.on_multi_animation_finished)
        else:
            settings = load_settings()
            immediate = settings.get("result_display_immediate", False)
            if not immediate and self.pending_draw_result is not None:
                self.all_draw_results.append(self.pending_draw_result)
                self.update_history_display()
                self.pending_draw_result = None
                self.flush_pending_logs()
                self.defer_logs = False
            self.current_draw_index += 1
            QTimer.singleShot(800, self.do_next_draw)

    def on_multi_animation_finished(self):
        pid = self.multi_results[self.current_multi_index]
        person = self.pm.get_person_by_id(pid)
        self.last_draw_person = person
        self.update_last_result_display()
        settings = load_settings()
        if not settings.get("result_display_immediate", False):
            self.current_result_label.setText(f"第{self.current_draw_index+1}次: {self.format_person_name(person, settings)}")
        # 多次抽取自动关闭：根据复选框决定是否传时间
        if settings.get("auto_close_multi_enabled", True):
            auto_close = settings.get("auto_close_multi_seconds", 0.5)
        else:
            auto_close = 0
        self.close_current_result_dialog()
        self.current_result_dialog = ResultDialog(person, self, auto_close)
        self.current_result_dialog.show()
        self.current_multi_index += 1
        QTimer.singleShot(500, self.show_next_multi_result)

    def start_animation(self, target_id, callback):
        if self.is_animating:
            return
        self.animation_target_id = target_id
        participant_ids = self.pm.get_participants()
        if not participant_ids:
            return
        seq = []
        for _ in range(5):
            shuffled = participant_ids.copy()
            random.shuffle(shuffled)
            seq.extend(shuffled)
        seq.extend([target_id] * 10)
        total_steps = len(seq)
        self.animation_sequence = seq
        self.animation_index = 0
        self.is_animating = True
        self.result_callback = callback
        for card in self.cards.values():
            card.set_highlight(False)

        settings = load_settings()
        speed_front = settings.get("anim_speed_front", 40)
        speed_mid = settings.get("anim_speed_mid", 80)
        speed_back = settings.get("anim_speed_back", 120)

        if total_steps <= 40:
            front_count = total_steps
            mid_count = 0
            back_count = 0
        else:
            front_count = total_steps - 40
            mid_count = 20
            back_count = 20
        self.animation_intervals = []
        for i in range(front_count):
            self.animation_intervals.append(speed_front)
        for i in range(mid_count):
            self.animation_intervals.append(speed_mid)
        for i in range(back_count):
            self.animation_intervals.append(speed_back)

        self.animation_timer.start(self.animation_intervals[0])

    def animate_step(self):
        if not self.is_animating:
            return
        if self.animation_index >= len(self.animation_sequence):
            self.animation_timer.stop()
            self.is_animating = False
            if self.animation_target_id in self.cards:
                self.cards[self.animation_target_id].set_highlight(True)
            if self.result_callback:
                self.result_callback()
            person = self.pm.get_person_by_id(self.animation_target_id)
            if person:
                settings = load_settings()
                show_nick = settings.get("last_show_nickname", True)
                show_real = settings.get("last_show_realname", False)
                if show_nick and show_real and person.realname:
                    name = f"{person.nickname} ({person.realname})"
                elif show_real and person.realname:
                    name = person.realname
                else:
                    name = person.nickname
                self.last_result_label.setText(f"上次结果: {name}")
            return
        for card in self.cards.values():
            card.set_highlight(False)
        current_id = self.animation_sequence[self.animation_index]
        if current_id in self.cards:
            self.cards[current_id].set_highlight(True)
        self.animation_index += 1
        if self.animation_index < len(self.animation_intervals):
            next_interval = self.animation_intervals[self.animation_index]
        else:
            next_interval = 100
        self.animation_timer.setInterval(next_interval)

    # ========== 窗口状态持久化 ==========
    def save_main_splitter_state(self):
        settings = load_settings()
        settings["main_splitter_sizes"] = self.main_splitter.sizes()
        save_settings(settings)

    def save_right_splitter_state(self):
        settings = load_settings()
        settings["right_splitter_sizes"] = self.right_splitter.sizes()
        save_settings(settings)

    def load_window_state(self):
        settings = load_settings()
        geometry = settings.get("window_geometry")
        if geometry:
            self.restoreGeometry(QByteArray.fromHex(geometry.encode()))
        else:
            self.resize(1200, 800)
            self.center_window()
        main_sizes = settings.get("main_splitter_sizes")
        if main_sizes and len(main_sizes) == 2:
            self.main_splitter.setSizes(main_sizes)
        right_sizes = settings.get("right_splitter_sizes")
        if right_sizes and len(right_sizes) == 2:
            self.right_splitter.setSizes(right_sizes)

    def center_window(self):
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)

    def closeEvent(self, event):
        settings = load_settings()
        settings["window_geometry"] = self.saveGeometry().toHex().data().decode()
        save_settings(settings)
        event.ignore()
        self.hide()
        self.tray_icon.showMessage("提示", "程序已最小化到系统托盘，双击图标可恢复。", QSystemTrayIcon.Information, 2000)

    # ========== 系统托盘 ==========
    def create_tray(self):
        self.tray_icon = QSystemTrayIcon(self)
        if os.path.exists(ICON_PATH):
            self.tray_icon.setIcon(QIcon(ICON_PATH))
        else:
            self.tray_icon.setIcon(self.style().standardIcon(QStyle.SP_ComputerIcon))
        self.tray_icon.setToolTip("随机抽选工具")
        menu = QMenu()
        show_action = menu.addAction("显示主界面")
        show_action.triggered.connect(self.show_and_activate)
        quit_action = menu.addAction("退出")
        quit_action.triggered.connect(self.quit_app)
        self.tray_icon.setContextMenu(menu)
        self.tray_icon.activated.connect(self.on_tray_activated)
        self.tray_icon.show()

    def on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_and_activate()

    def show_and_activate(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()

    def quit_app(self):
        self.tray_icon.hide()
        QApplication.quit()

    # ========== 单实例 ==========
    def start_local_server(self):
        self.server = QLocalServer(self)
        server_name = "App14345_single_instance"
        self.server.listen(server_name)
        self.server.newConnection.connect(self.handle_client_connection)

    def handle_client_connection(self):
        socket = self.server.nextPendingConnection()
        if socket:
            socket.waitForReadyRead(100)
            data = socket.readAll().data()
            if data == b"show":
                self.show_and_activate()
            socket.disconnectFromServer()

# ==================== 辅助函数 ====================
def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "last_photo_dir": os.path.expanduser("~"),
        "card_show_nickname": True,
        "card_show_realname": False,
        "last_show_nickname": True,
        "last_show_realname": False,
        "draw_show_nickname": True,
        "draw_show_realname": False,
        "log_panel_visible": True,
        "draw_k": 1,
        "draw_repeat": 1,
        "show_title_copyright": True,
        "auto_close_single_enabled": False,
        "auto_close_single_seconds": 2.0,
        "auto_close_multi_enabled": True,
        "auto_close_multi_seconds": 0.5,
        "anim_speed_front": 40,
        "anim_speed_mid": 80,
        "anim_speed_back": 120,
        "result_display_immediate": False,
        "show_history_list": True,
        "show_process_log": True,
        "main_splitter_sizes": [800, 400],
        "right_splitter_sizes": [300, 400]
    }

def save_settings(settings):
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, indent=2)

def load_stats():
    if os.path.exists(STATS_FILE):
        with open(STATS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def record_draw_result(chosen_ids):
    stats = load_stats()
    stats.append({
        "time": datetime.now().isoformat(),
        "chosen_ids": chosen_ids
    })
    with open(STATS_FILE, 'w', encoding='utf-8') as f:
        json.dump(stats, f, indent=2)

def single_instance_check():
    socket = QLocalSocket()
    socket.connectToServer("App14345_single_instance")
    if socket.waitForConnected(500):
        socket.write(b"show")
        socket.flush()
        socket.waitForBytesWritten(500)
        socket.disconnectFromServer()
        return False
    else:
        return True

def main():
    #ensure_default_icon()
    if not single_instance_check():
        sys.exit(0)
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(ICON_PATH) if os.path.exists(ICON_PATH) else QIcon())
    person_manager = PersonManager()
    logger = Logger()
    picker = RandomPicker(logger, person_manager=person_manager)
    window = MainWindow(person_manager, picker, logger)
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
